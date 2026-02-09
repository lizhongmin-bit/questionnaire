from io import BytesIO
from typing import Iterable, List, Tuple
from openpyxl import Workbook, load_workbook


def parse_id_xlsx(content: bytes, expected_header: str = "ID") -> Tuple[List[str], List[str]]:
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["空文件"]
    header = str(rows[0][0]).strip() if rows[0][0] is not None else ""
    if header != expected_header:
        return [], [f"首列表头必须为{expected_header}"]
    ids = []
    errors = []
    seen = set()
    for row in rows[1:]:
        if not row:
            continue
        value = row[0]
        if value is None:
            continue
        real_id = str(value).strip()
        if not real_id:
            continue
        if real_id in seen:
            continue
        seen.add(real_id)
        ids.append(real_id)
    return ids, errors


def build_id_template(expected_header: str = "ID") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    ws.append([expected_header])
    ws.append(["示例: 13800000000"])  # 说明行
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_links_xlsx(rows: Iterable[Tuple[str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "问卷链接"
    ws.append(["ID", "问卷链接"])
    for real_id, link in rows:
        ws.append([real_id, link])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_links_txt(rows: Iterable[Tuple[str, str]]) -> bytes:
    lines = []
    for real_id, link in rows:
        lines.append(f"{real_id},{link}")
    return "\n".join(lines).encode("utf-8")


def build_answers_xlsx(headers: List[str], rows: List[List[str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "答案导出"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
